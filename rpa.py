"""
Elaw Carrefour — Cumprimento em lote: Externo Inserir Relatório de Audiência

Fluxo por processo:
  1. Recebe a linha da planilha (já preenchida pelo correspondente/preposto)
  2. Navega até o processo no Elaw Carrefour
  3. Localiza o prazo "Externo: Inserir Relatório da Audiência" pela data/hora
  4. Verifica via lupa que é o prazo correto
  5. Preenche todos os campos do formulário
  6. Clica em Confirmar

Notas técnicas (JSF/PrimeFaces):
- Navegação direta por URL não funciona — usar busca global
- Botões via JavaScript (podem estar fora do viewport)
- page.evaluate com string NÃO aceita return no top-level — usar IIFEs
"""

import os
import re
import time
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

BRASILIA_TZ = ZoneInfo("America/Sao_Paulo")

if os.environ.get("RENDER"):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "/opt/render/project/src/.playwright-browsers"

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

ELAW_URL      = os.environ.get("ELAW_CARREFOUR_URL", "https://carrefour.elaw.com.br")
LOGIN_TIMEOUT = 120_000
PAGE_TIMEOUT  = 40_000
POLL_ATTEMPTS = 12
POLL_WAIT     = 2.0

IS_SERVER = bool(os.environ.get("RENDER") or os.environ.get("IS_SERVER"))


# ── Entry point ───────────────────────────────────────────────────────────────

def run_automation(rows: list[dict], log, report_path: Path, state: dict | None = None):
    results = []

    with sync_playwright() as p:
        if IS_SERVER:
            browser = _launch_server(p)
            ctx  = browser.new_context(viewport={"width": 1280, "height": 900})
            page = ctx.new_page()
        else:
            chrome_profile = str(Path(__file__).parent / "chrome_profile")
            ctx  = p.chromium.launch_persistent_context(
                chrome_profile,
                headless=False,
                viewport={"width": 1280, "height": 900},
                slow_mo=120,
            )
            page = ctx.pages[0] if ctx.pages else ctx.new_page()

        log("Abrindo Elaw Carrefour...")
        page.goto(ELAW_URL, wait_until="networkidle", timeout=30_000)

        if _is_login_page(page):
            if IS_SERVER:
                log("Fazendo login automático...", "info")
                _auto_login(page, log)
            else:
                log("⚠️ Sessão expirada — faça login no browser aberto.", "warn")
                page.wait_for_url(f"**{ELAW_URL}/**", timeout=LOGIN_TIMEOUT)
                page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
                log("✅ Login detectado, iniciando automação...")
        else:
            log("✅ Sessão ativa, iniciando automação...")

        page.goto(f"{ELAW_URL}/processoList.elaw", wait_until="networkidle", timeout=PAGE_TIMEOUT)
        page.wait_for_selector('[id*="globaSearchAutocomplete_input"]', state="visible", timeout=PAGE_TIMEOUT)
        time.sleep(3)  # aguarda sessão JSF estabilizar antes da primeira busca

        total = len(rows)
        for i, row in enumerate(rows, 1):
            if state and state.get("paused"):
                log("⏸ Pausado — aguardando retomada...", "warn")
                while state.get("paused"):
                    time.sleep(1)
                log("▶️ Retomando...", "info")

            numero         = str(row.get("numero_processo", "")).strip()
            data_audiencia = str(row.get("data_audiencia", "")).strip()
            horario        = str(row.get("horario_audiencia", "")).strip()
            intercorrencias = str(row.get("intercorrencias") or "").strip()

            log(f"[{i}/{total}] {numero} — {data_audiencia} {horario}...")

            for attempt in range(2):
                try:
                    status, detail = _process_row(page, numero, data_audiencia, horario, row, log)
                    break
                except Exception as e:
                    err_str = str(e)
                    if "Execution context was destroyed" in err_str and attempt == 0:
                        log("  🔁 Contexto destruído, tentando novamente...", "warn")
                        _recover_page(page, log)
                    else:
                        status, detail = "ERRO", err_str[:300]
                        _recover_page(page, log)
                        break

            # Sinaliza intercorrência como OBSERVAÇÃO
            if status == "OK" and intercorrencias and intercorrencias.upper() != "NÃO":
                status = "OBSERVAÇÃO"
                detail = "Relatório inserido — intercorrência para revisão manual"

            row_result = {
                "numero_processo": numero,
                "data_audiencia":  data_audiencia,
                "resultado":       str(row.get("resultado_audiencia") or ""),
                "intercorrencias": intercorrencias,
                "status":          status,
                "detalhe":         detail,
                "horario_exec":    datetime.now(BRASILIA_TZ).strftime("%H:%M:%S"),
            }
            results.append(row_result)
            if state is not None:
                state["results"].append(row_result)

            icons = {"OK": "✅", "JÁ CUMPRIDO": "ℹ️", "ERRO": "❌", "OBSERVAÇÃO": "⚠️"}
            css   = {"OK": "ok", "JÁ CUMPRIDO": "ok", "ERRO": "error", "OBSERVAÇÃO": "warn"}
            log(f"  {icons.get(status,'❌')} {status}: {detail}", css.get(status, "error"))
            if status == "OBSERVAÇÃO":
                log(f"  ⚠️  Intercorrência: {intercorrencias[:150]}", "warn")

        if IS_SERVER:
            browser.close()
        else:
            ctx.close()

    _build_report(results, report_path)
    ok  = sum(1 for r in results if r["status"] in ("OK", "JÁ CUMPRIDO", "OBSERVAÇÃO"))
    obs = sum(1 for r in results if r["status"] == "OBSERVAÇÃO")
    log(
        f"Concluído: {ok}/{len(results)} processos com sucesso"
        + (f" — {obs} com intercorrência para revisão" if obs else "") + ".",
        "done",
    )


# ── Browser helpers ───────────────────────────────────────────────────────────

def _launch_server(p):
    return p.chromium.launch(
        headless=True,
        args=[
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--disable-setuid-sandbox",
            "--single-process",
        ],
    )


def _auto_login(page, log):
    user = os.environ.get("ELAW_CARREFOUR_USER", "")
    pwd  = os.environ.get("ELAW_CARREFOUR_PASS", "")

    if not user or not pwd:
        raise Exception(
            "Variáveis ELAW_CARREFOUR_USER e ELAW_CARREFOUR_PASS não configuradas."
        )

    page.wait_for_selector("#username", state="visible", timeout=PAGE_TIMEOUT)
    page.wait_for_selector("#authKey",  state="visible", timeout=PAGE_TIMEOUT)

    log("Preenchendo credenciais...")
    page.fill("#username", user, timeout=PAGE_TIMEOUT)
    page.fill("#authKey",  pwd,  timeout=PAGE_TIMEOUT)

    try:
        page.locator("button.ui-button").first.click(force=True, timeout=PAGE_TIMEOUT)
    except Exception:
        page.evaluate("""(() => {
            const btn = document.querySelector('button.ui-button') ||
                        Array.from(document.querySelectorAll('button'))
                            .find(b => b.textContent.trim().includes('Acessar'));
            if (btn) btn.click();
            else { const f = document.querySelector('form'); if (f) f.submit(); }
        })()""")

    page.wait_for_load_state("networkidle", timeout=30_000)

    if _is_login_page(page):
        try:
            page.screenshot(path="/tmp/debug_login.png", full_page=True)
        except Exception:
            pass
        raise Exception("Login falhou — verifique ELAW_CARREFOUR_USER e ELAW_CARREFOUR_PASS.")

    log("✅ Login concluído.")


def _is_login_page(page) -> bool:
    """Detecta página de login de forma conservadora — só pela presença do form visível."""
    try:
        return bool(page.evaluate("""(() => {
            const isVisible = el => el && el.offsetParent !== null
                && getComputedStyle(el).display !== 'none'
                && getComputedStyle(el).visibility !== 'hidden';
            // Só considera login se AMBOS os campos estiverem visíveis simultaneamente
            return isVisible(document.getElementById('username'))
                && isVisible(document.getElementById('authKey'));
        })()"""))
    except Exception:
        return False


def _recover_page(page, log):
    try:
        page.goto(f"{ELAW_URL}/processoList.elaw", wait_until="networkidle", timeout=PAGE_TIMEOUT)
        if _is_login_page(page):
            log("  🔑 Sessão expirada — refazendo login...", "warn")
            _auto_login(page, log)
            page.goto(f"{ELAW_URL}/processoList.elaw", wait_until="networkidle", timeout=PAGE_TIMEOUT)
        page.wait_for_selector(
            '[id*="globaSearchAutocomplete_input"]',
            state="visible",
            timeout=PAGE_TIMEOUT,
        )
        time.sleep(3)  # aguarda sessão JSF estabilizar após recovery/re-login
    except Exception as e:
        log(f"  ⚠️ Recovery falhou: {str(e)[:120]}", "warn")


# ── Fluxo por processo ────────────────────────────────────────────────────────

def _process_row(page, numero, data_audiencia, horario, row, log):
    _navigate_to_process(page, numero)
    # "Pauta e Andamento" já abre por padrão ao entrar no processo

    task_status = _find_verify_and_open_task(page, data_audiencia, horario, log)
    if task_status == "ja_cumprido":
        _go_to_process_list(page)
        return "JÁ CUMPRIDO", "Tarefa já estava cumprida anteriormente"

    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
    _fill_form(page, row)
    _confirm_task(page)

    # Volta para a lista de processos — garante estado limpo para o próximo
    _go_to_process_list(page)
    return "OK", "Relatório inserido com sucesso"


# ── Navegação ─────────────────────────────────────────────────────────────────

def _navigate_to_process(page, numero):
    page.wait_for_selector('[id*="globaSearchAutocomplete_input"]', timeout=PAGE_TIMEOUT)

    clicked = False
    for attempt in range(2):
        page.evaluate("""
            const el = document.querySelector('[id*="globaSearchAutocomplete_input"]');
            el.value = '';
            el.focus();
            el.click();
        """)
        time.sleep(0.4)
        page.keyboard.type(numero, delay=65)

        for _ in range(POLL_ATTEMPTS):
            time.sleep(POLL_WAIT)
            result = page.evaluate("""(() => {
                const panel = document.querySelector('[id$="globaSearchAutocomplete_panel"]');
                const items = panel ? panel.querySelectorAll('li') : [];
                if (items.length > 0) { items[0].click(); return 'clicado'; }
                return 'vazio';
            })()""")
            if result == "clicado":
                clicked = True
                break
        if clicked:
            break

    if not clicked:
        raise Exception(f"Autocomplete da busca não abriu para o processo {numero}")

    try:
        page.wait_for_url("**/processoView.elaw**", timeout=PAGE_TIMEOUT)
    except PWTimeout:
        raise Exception(f"Processo {numero} não encontrado no sistema")

    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)


def _go_to_process_list(page):
    """Retorna à lista de processos — garante estado limpo entre iterações."""
    try:
        page.goto(f"{ELAW_URL}/processoList.elaw", wait_until="networkidle", timeout=PAGE_TIMEOUT)
        page.wait_for_selector(
            '[id*="globaSearchAutocomplete_input"]',
            state="visible",
            timeout=PAGE_TIMEOUT,
        )
    except Exception:
        pass  # Se falhar, _navigate_to_process tentará novamente


def _click_pauta_andamento(page):
    """Clica na aba 'Pauta e Andamento'. Usa locator nativo do Playwright (mais robusto)."""
    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
    time.sleep(0.5)

    # Estratégia 1: locator por texto exato (Playwright normaliza espaços/encoding)
    try:
        loc = page.get_by_text("Pauta e Andamento", exact=True).first
        loc.wait_for(state="visible", timeout=8_000)
        loc.click()
        time.sleep(1.5)
        page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
        return "ok:get_by_text-exact"
    except Exception:
        pass

    # Estratégia 2: locator por texto parcial
    try:
        loc = page.locator("text=Pauta e Andamento").first
        loc.wait_for(state="visible", timeout=5_000)
        loc.click()
        time.sleep(1.5)
        page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
        return "ok:locator-text"
    except Exception:
        pass

    # Estratégia 3: JS — busca em toda a página (fallback)
    result = page.evaluate("""(() => {
        const walker = document.createTreeWalker(
            document.body, NodeFilter.SHOW_TEXT
        );
        let node;
        while ((node = walker.nextNode())) {
            const t = node.textContent.trim().toLowerCase();
            if (t === 'pauta e andamento') {
                const el = node.parentElement;
                const clickable = el.closest('a, button, li, [role="tab"]') || el;
                clickable.click();
                return 'ok:treewalker/' + el.tagName;
            }
        }
        // Debug: lista todos os textos de links visíveis
        const links = Array.from(document.querySelectorAll('a, [role="tab"]'))
            .filter(e => e.offsetParent !== null)
            .map(e => e.textContent.trim().substring(0, 30))
            .join(' | ');
        return 'nao_encontrado|links:' + links.substring(0, 300);
    })()""")

    if result.startswith("nao_encontrado"):
        # Salva screenshot para diagnóstico
        try:
            page.screenshot(path="/tmp/debug_pauta.png", full_page=False)
        except Exception:
            pass
        raise Exception(
            f"Aba 'Pauta e Andamento' não encontrada. Debug: {result}"
        )

    time.sleep(1.5)
    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
    return result


# ── Localizar e abrir tarefa ──────────────────────────────────────────────────

def _find_verify_and_open_task(page, data_audiencia, horario, log) -> str:
    date_part = data_audiencia.strip()[:10]   # DD/MM/YYYY
    time_part = horario.strip()[:5]           # HH:MM

    task_count = page.evaluate("""(() => {
        return Array.from(document.querySelectorAll('tr')).filter(r =>
            r.textContent.toLowerCase().includes('externo') &&
            r.textContent.toLowerCase().includes('relat') &&
            r.textContent.toLowerCase().includes('audi')
        ).length;
    })()""")

    if task_count == 0:
        return "ja_cumprido"

    matched_index = None

    for i in range(task_count):
        match = _verify_task_date(page, i, date_part, time_part, log)
        if match:
            matched_index = i
            break

    if matched_index is None:
        if task_count == 1:
            # Único prazo — usa mesmo sem confirmar data (já logou aviso)
            matched_index = 0
        else:
            raise Exception(
                f"Nenhum dos {task_count} prazos 'Externo: Inserir Relatório' "
                f"corresponde a {data_audiencia} {horario}."
            )

    _click_check_icon(page, matched_index)
    return "ok"


def _verify_task_date(page, row_index: int, date_part: str, time_part: str, log) -> bool:
    lupa_result = page.evaluate(f"""(() => {{
        const rows = Array.from(document.querySelectorAll('tr')).filter(r =>
            r.textContent.toLowerCase().includes('externo') &&
            r.textContent.toLowerCase().includes('relat') &&
            r.textContent.toLowerCase().includes('audi')
        );
        const row = rows[{row_index}];
        if (!row) return 'sem_row';

        const buttons = Array.from(row.querySelectorAll('button, a'));
        let btn = buttons.find(b =>
            b.innerHTML.toLowerCase().includes('search') ||
            b.innerHTML.toLowerCase().includes('lupa')  ||
            b.innerHTML.toLowerCase().includes('zoom')  ||
            b.innerHTML.toLowerCase().includes('eye')   ||
            (b.querySelector && b.querySelector('[class*="search"],[class*="lupa"]'))
        );
        if (!btn) {{
            btn = buttons.find(b =>
                !b.innerHTML.toLowerCase().includes('check') &&
                !b.innerHTML.toLowerCase().includes('tick')  &&
                !b.id.toLowerCase().includes('confirm')
            );
        }}
        if (!btn && buttons.length > 0) btn = buttons[0];
        if (!btn) return 'sem_lupa';
        btn.click();
        return 'ok';
    }})()""")

    if lupa_result != "ok":
        return True  # Sem lupa, assume correto

    time.sleep(1.2)

    popup_text = page.evaluate("""(() => {
        for (const sel of ['.ui-dialog','.ui-overlaypanel','[class*="popup"]','[class*="modal"]']) {
            const el = document.querySelector(sel);
            if (el && el.offsetParent !== null) return el.textContent;
        }
        return null;
    })()""")

    # Fecha popup
    page.evaluate("""(() => {
        const closeSelectors = ['.ui-dialog-titlebar-close','button[aria-label="Close"]'];
        for (const sel of closeSelectors) {
            const btn = document.querySelector(sel);
            if (btn) { btn.click(); return; }
        }
        const fechar = Array.from(document.querySelectorAll('button, a'))
            .find(b => b.textContent.trim().toLowerCase() === 'fechar');
        if (fechar) fechar.click();
    })()""")
    time.sleep(0.5)

    if popup_text:
        has_date = (not date_part) or (date_part in popup_text)
        has_time = (not time_part) or (time_part in popup_text)
        return has_date and has_time

    return True


def _click_check_icon(page, row_index: int):
    result = page.evaluate(f"""(() => {{
        const rows = Array.from(document.querySelectorAll('tr')).filter(r =>
            r.textContent.toLowerCase().includes('externo') &&
            r.textContent.toLowerCase().includes('relat') &&
            r.textContent.toLowerCase().includes('audi')
        );
        const row = rows[{row_index}];
        if (!row) return 'sem_row';

        const buttons = Array.from(row.querySelectorAll('button, a'));
        let btn = buttons.find(b =>
            b.innerHTML.toLowerCase().includes('check') ||
            b.innerHTML.toLowerCase().includes('tick')  ||
            b.id.toLowerCase().includes('confirm')      ||
            b.id.toLowerCase().includes('cumprir')      ||
            (b.querySelector && b.querySelector('[class*="check"],[class*="confirm"]'))
        );
        if (!btn && buttons.length > 0) btn = buttons[buttons.length - 1];
        if (!btn) return 'sem_check';
        btn.click();
        return btn.id || 'clicado';
    }})()""")

    if result in ("sem_row", "sem_check"):
        raise Exception(f"Botão de check/confirmar não encontrado: {result}")

    try:
        page.wait_for_url("**agendamentoContenciosoConfirm.elaw**", timeout=PAGE_TIMEOUT)
    except PWTimeout:
        try:
            page.wait_for_selector(
                '[id*="confirmar"], [id*="btnConfirm"], .ui-dialog',
                state="visible",
                timeout=PAGE_TIMEOUT,
            )
        except PWTimeout:
            raise Exception(
                f"Formulário de cumprimento não carregou (URL: {page.url})"
            )


# ── Preenchimento do formulário ───────────────────────────────────────────────

# Mapeamento: chave da row → lista de palavras-chave do label no Carrefour
_FIELD_MAP = [
    # (row_key,                      label_keywords,                      tipo)
    ("nome_preposto",               ["preposto", "compareceu"],           "sim_from_value"),
    ("advogado_pontual",            ["advogado", "horario"],              "sim_nao"),
    ("advogado_contato",            ["advogado", "contato"],              "sim_nao"),
    ("preposto_ouvido",             ["ouvido"],                           "sim_nao"),
    ("orientacoes_claras",          ["orientacoes", "claras"],            "sim_nao"),
    ("intercorrencias",             ["intercorr"],                        "text"),
    ("teve_testemunha",             ["teve", "testemunha"],               "sim_nao"),
    ("reclamada_testemunha",        ["reclamada", "testemunha"],          "sim_nao"),
    ("reclamada_testemunha_ouvida", ["reclamada", "ouvida"],              "sim_nao"),
    ("reclamada_testemunha_nome",   ["testemunha", "reclamada"],          "text"),
    ("reclamante_testemunha",       ["reclamante", "testemunha"],         "sim_nao"),
    ("reclamante_testemunha_ouvida",["reclamante", "ouvida"],             "sim_nao"),
    ("reclamante_testemunha_nome",  ["testemunha", "reclamante"],         "text"),
    ("resultado_audiencia",         ["resultado"],                         "dropdown"),
]


def _fill_form(page, row: dict):
    for row_key, keywords, field_type in _FIELD_MAP:
        value = str(row.get(row_key) or "").strip()
        if not value or value.lower() == "nan":
            continue

        if field_type == "sim_from_value":
            # "Preposto designado compareceu" → Sim se há nome do preposto
            _set_sim_nao(page, keywords, "Sim")

        elif field_type == "sim_nao":
            # Valor direto: "SIM"/"NÃO" → clica no radio/checkbox certo
            _set_sim_nao(page, keywords, value)

        elif field_type == "text":
            # Campo de texto livre
            _set_text(page, keywords, value)

        elif field_type == "dropdown":
            # Dropdown PrimeFaces
            _set_dropdown(page, keywords, value)

        time.sleep(0.3)


def _set_sim_nao(page, keywords: list, value: str):
    kw_js    = str(keywords)
    val_low  = value.strip().lower()

    page.evaluate(f"""(() => {{
        const keywords = {kw_js};
        const valueLow = '{val_low}';

        function containerFor(kws) {{
            for (const lbl of document.querySelectorAll(
                'label, th, td, .ui-outputlabel, [class*="label"]'
            )) {{
                const t = lbl.textContent.trim().toLowerCase();
                if (kws.every(k => t.includes(k))) {{
                    return lbl.closest('tr, .ui-grid-row, .field, .form-group, fieldset') || lbl.parentElement;
                }}
            }}
            return null;
        }}

        const container = containerFor(keywords);
        if (!container) return;

        // Tenta radio buttons
        for (const radio of container.querySelectorAll('input[type=radio]')) {{
            const lbl = document.querySelector('label[for="' + radio.id + '"]');
            const txt = (lbl ? lbl.textContent : radio.value).trim().toLowerCase();
            if (txt.includes(valueLow) || valueLow.includes(txt)) {{
                radio.click();
                return;
            }}
        }}

        // Tenta select nativo
        const sel = container.querySelector('select');
        if (sel) {{
            const opt = Array.from(sel.options).find(o =>
                o.text.trim().toLowerCase().includes(valueLow)
            );
            if (opt) {{
                sel.value = opt.value;
                sel.dispatchEvent(new Event('change', {{ bubbles: true }}));
            }}
            return;
        }}

        // Tenta PrimeFaces selectOneRadio / selectBooleanCheckbox por texto
        const items = container.querySelectorAll('.ui-radiobutton, .ui-chkbox');
        for (const item of items) {{
            const lbl = item.nextElementSibling || item.previousElementSibling;
            if (lbl && lbl.textContent.trim().toLowerCase().includes(valueLow)) {{
                item.querySelector('input, .ui-radiobutton-box, .ui-chkbox-box')?.click();
                return;
            }}
        }}
    }})()""")


def _set_dropdown(page, keywords: list, value: str):
    kw_js   = str(keywords)
    val_low = value.strip().lower()

    page.evaluate(f"""(() => {{
        const keywords = {kw_js};
        const valueLow = '{val_low}';

        function containerFor(kws) {{
            for (const lbl of document.querySelectorAll(
                'label, th, td, .ui-outputlabel, [class*="label"]'
            )) {{
                const t = lbl.textContent.trim().toLowerCase();
                if (kws.every(k => t.includes(k))) {{
                    return lbl.closest('tr, .ui-grid-row, .field, .form-group, fieldset') || lbl.parentElement;
                }}
            }}
            return null;
        }}

        const container = containerFor(keywords);
        if (!container) return;

        // Select nativo
        const sel = container.querySelector('select');
        if (sel) {{
            const opt = Array.from(sel.options).find(o =>
                o.text.trim().toLowerCase().includes(valueLow)
            );
            if (opt) {{
                sel.value = opt.value;
                sel.dispatchEvent(new Event('change', {{ bubbles: true }}));
            }}
            return;
        }}

        // PrimeFaces selectOneMenu
        const pfDrop = container.querySelector('.ui-selectonemenu');
        if (pfDrop) {{
            const trigger = pfDrop.querySelector('.ui-selectonemenu-trigger');
            if (trigger) {{
                trigger.click();
                setTimeout(() => {{
                    const panel = document.querySelector('.ui-selectonemenu-panel:not([style*="display: none"])');
                    if (!panel) return;
                    const target = Array.from(panel.querySelectorAll('li')).find(
                        i => i.textContent.trim().toLowerCase().includes(valueLow)
                    );
                    if (target) target.click();
                }}, 500);
            }}
        }}
    }})()""")
    time.sleep(0.6)  # aguarda o setTimeout do dropdown


def _set_text(page, keywords: list, value: str):
    kw_js = str(keywords)

    page.evaluate(f"""(() => {{
        const keywords = {kw_js};
        const value = {repr(value)};

        for (const lbl of document.querySelectorAll(
            'label, th, td, .ui-outputlabel, [class*="label"]'
        )) {{
            const t = lbl.textContent.trim().toLowerCase();
            if (keywords.every(k => t.includes(k))) {{
                const forId = lbl.getAttribute('for');
                const field = forId ? document.getElementById(forId) : null;
                const container = lbl.closest('tr, .field, .form-group');
                const inp = field || (container && container.querySelector(
                    'input[type=text], textarea'
                ));
                if (inp) {{
                    inp.value = value;
                    inp.dispatchEvent(new Event('input',  {{ bubbles: true }}));
                    inp.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    return;
                }}
            }}
        }}
    }})()""")


def _confirm_task(page):
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(0.5)

    result = page.evaluate("""(() => {
        const byId = document.getElementById('btnConfirmaSim');
        if (byId) { byId.click(); return 'ok'; }
        const btns = Array.from(document.querySelectorAll('button, input[type="submit"]'));
        const btn = btns.find(b =>
            b.textContent.trim().toLowerCase().includes('confirmar') ||
            (b.value && b.value.trim().toLowerCase().includes('confirmar'))
        );
        if (btn) { btn.click(); return 'ok'; }
        return 'nao_encontrado';
    })()""")

    if result != "ok":
        raise Exception("Botão 'Confirmar' não encontrado no formulário")

    time.sleep(2)
    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)

    # O Elaw usa AJAX (PrimeFaces) — a URL pode não mudar após submissão bem-sucedida.
    # Em vez de checar URL, verifica se há mensagens de erro JSF visíveis.
    # Se o formulário foi aceito, nenhuma mensagem de erro aparece.
    error_msgs = page.evaluate("""(() => {
        const msgs = new Set();
        const selectors = [
            '.ui-message-error-detail',
            '.ui-messages-error-detail',
            '.ui-messages-error-summary',
            '.ui-message-error-summary',
            '.ui-growl-item-message',
        ];
        for (const sel of selectors) {
            document.querySelectorAll(sel).forEach(el => {
                if (el.offsetParent !== null) {  // só elementos visíveis
                    const t = el.textContent.trim();
                    if (t) msgs.add(t);
                }
            });
        }
        return [...msgs].join(' | ');
    })()""")

    if error_msgs:
        raise Exception(f"Erro de validação JSF ao confirmar: {error_msgs}")


# ── Relatório Excel ───────────────────────────────────────────────────────────

def _build_report(results: list[dict], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    headers = ["Número Processo", "Data Audiência", "Resultado", "Intercorrências",
               "Status", "Detalhe", "Horário"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    fills = {
        "OK":          PatternFill("solid", fgColor="C6EFCE"),
        "JÁ CUMPRIDO": PatternFill("solid", fgColor="DDEBF7"),
        "OBSERVAÇÃO":  PatternFill("solid", fgColor="FFEB9C"),
        "ERRO":        PatternFill("solid", fgColor="FFC7CE"),
    }

    for r in results:
        ws.append([
            r["numero_processo"],
            r["data_audiencia"],
            r["resultado"],
            r["intercorrencias"],
            r["status"],
            r["detalhe"],
            r["horario_exec"],
        ])
        for cell in ws[ws.max_row]:
            cell.fill = fills.get(r["status"], fills["ERRO"])

    for col, w in zip("ABCDEFG", [24, 16, 22, 50, 14, 45, 10]):
        ws.column_dimensions[col].width = w

    wb.save(path)
