import json
import os
import queue
import re
import threading
import traceback
import unicodedata
import uuid
from functools import wraps
from pathlib import Path

from flask import (Flask, jsonify, redirect, render_template,
                   request, Response, send_file, session, url_for)
from openpyxl import load_workbook

import rpa as rpa_module

# ── Normalização ──────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Lowercase + remove acentos + remove ?, *, \t, (, ) e espaços duplos."""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r'[\t\*\?\(\)]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ── Mapeamento de colunas ─────────────────────────────────────────────────────

# Chave: header normalizado (sem acento, sem ?, *, tab)
# Valor: nome interno usado em row dict / rpa.py
_COL_MAP = {
    # Identificação
    "numero do processo":                                           "numero_processo",
    "id processo":                                                  "id_processo",
    "npc":                                                          "npc",
    "nome do reclamante":                                           "nome_reclamante",
    # Audiência
    "data da audiencia":                                            "data_audiencia",
    "horario da audiencia":                                         "horario_audiencia",
    # Preposto / advogado
    "nome preposto":                                                "nome_preposto",
    "nome do advogado que o acompanhou":                            "nome_advogado",
    "o advogado chegou no horario da audiencia":                    "advogado_pontual",
    "o advogado entrou em contato para orientacoes antes da audiencia": "advogado_contato",
    # Perguntas do preposto
    "voce foi ouvido":                                              "preposto_ouvido",
    "as orientacoes recebidas foram claras e suficientes":          "orientacoes_claras",
    # Intercorrências (header longo — usa prefixo)
    "houve alguma intercorrencias durante a audiencia":             "intercorrencias",
    # Testemunhas — Reclamada
    "teve testemunha":                                              "teve_testemunha",
    "reclamada levou testemunha":                                   "reclamada_testemunha",
    "se sim testemunha foi ouvida":                                 "reclamada_testemunha_ouvida",
    "nome da testemunha reclamada":                                 "reclamada_testemunha_nome",
    # Testemunhas — Reclamante
    "reclamante levou testemunha":                                  "reclamante_testemunha",
    "se sim testemunha foi ouvida reclamante":                      "reclamante_testemunha_ouvida",
    "nome da testemunha reclamante":                                "reclamante_testemunha_nome",
    # Resultado
    "externo - resultado da audiencia":                             "resultado_audiencia",
    "externo resultado da audiencia":                               "resultado_audiencia",
}

def _map_header(raw: str) -> str | None:
    """Mapeia header da planilha para chave interna."""
    norm = _norm(raw or "")
    if norm in _COL_MAP:
        return _COL_MAP[norm]
    # Prefixo (para headers longos como o de intercorrências)
    for key, val in _COL_MAP.items():
        if norm.startswith(key):
            return val
    return None


# ── App setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-troque-em-producao")

UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "uploads"))
OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", "output"))
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

APP_USER = os.environ.get("APP_USER", "admin")
APP_PASS = os.environ.get("APP_PASS", "admin")

_log_queue: queue.Queue = queue.Queue()
_state = {
    "running":      False,
    "done":         False,
    "report_ready": False,
    "session_id":   None,
    "paused":       False,
    "results":      [],
}


# ── Auth ──────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if (request.form.get("user") == APP_USER and
                request.form.get("pass") == APP_PASS):
            session["logged_in"] = True
            return redirect(url_for("index"))
        error = "Usuário ou senha incorretos."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        return jsonify({"erro": "Envie um arquivo .xlsx"}), 400

    session_id = uuid.uuid4().hex
    upload_dir = UPLOAD_DIR / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    dest = upload_dir / "planilha.xlsx"
    file.save(dest)

    try:
        rows = _parse_spreadsheet(dest)
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    session["upload_session"] = session_id
    preview = [
        {
            "processo": r["numero_processo"],
            "data":     str(r.get("data_audiencia") or ""),
            "horario":  str(r.get("horario_audiencia") or ""),
            "resultado": str(r.get("resultado_audiencia") or ""),
        }
        for r in rows[:5]
    ]
    return jsonify({"ok": True, "total": len(rows), "preview": preview})


@app.route("/execute", methods=["POST"])
@login_required
def execute():
    if _state["running"]:
        return jsonify({"erro": "Já há uma execução em andamento."}), 400

    sid = session.get("upload_session")
    if not sid:
        return jsonify({"erro": "Faça upload da planilha primeiro."}), 400

    dest = UPLOAD_DIR / sid / "planilha.xlsx"
    if not dest.exists():
        return jsonify({"erro": "Arquivo não encontrado. Faça upload novamente."}), 400

    try:
        rows = _parse_spreadsheet(dest)
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    out_dir = OUTPUT_DIR / sid
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "relatorio_audiencias.xlsx"

    while not _log_queue.empty():
        _log_queue.get_nowait()
    _state.update({"running": True, "done": False, "report_ready": False,
                   "session_id": sid, "paused": False, "results": []})

    thread = threading.Thread(target=_run, args=(rows, report_path), daemon=True)
    thread.start()
    return jsonify({"ok": True, "total": len(rows)})


@app.route("/logs")
@login_required
def logs():
    def generate():
        while True:
            try:
                msg = _log_queue.get(timeout=8)
                yield f"data: {msg}\n\n"
                if json.loads(msg).get("done"):
                    break
            except queue.Empty:
                yield 'data: {"heartbeat":true}\n\n'

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/status")
@login_required
def status():
    return jsonify({k: v for k, v in _state.items() if k != "results"})


@app.route("/pause", methods=["POST"])
@login_required
def pause():
    _state["paused"] = True
    return jsonify({"ok": True})


@app.route("/resume", methods=["POST"])
@login_required
def resume():
    _state["paused"] = False
    return jsonify({"ok": True})


@app.route("/partial-report")
@login_required
def partial_report():
    results = _state.get("results", [])
    if not results:
        return jsonify({"erro": "Nenhum resultado disponível ainda"}), 404
    path = Path("/tmp/relatorio_audiencias_parcial.xlsx")
    rpa_module._build_report(results, path)
    return send_file(path, as_attachment=True,
                     download_name="relatorio_audiencias_parcial.xlsx")


@app.route("/download")
@login_required
def download():
    sid = _state.get("session_id") or session.get("upload_session")
    if not sid:
        return jsonify({"erro": "Relatório não disponível"}), 404
    path = OUTPUT_DIR / sid / "relatorio_audiencias.xlsx"
    if not path.exists():
        return jsonify({"erro": "Relatório ainda não gerado"}), 404
    return send_file(path, as_attachment=True,
                     download_name="relatorio_audiencias.xlsx")


@app.route("/debug-screenshot")
@login_required
def debug_screenshot():
    path = Path("/tmp/debug_login.png")
    if not path.exists():
        return jsonify({"erro": "Nenhum screenshot disponível ainda"}), 404
    return send_file(path, mimetype="image/png")


# ── Error handlers ────────────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({"erro": "Rota não encontrada"}), 404


@app.errorhandler(500)
def internal_error(e):
    return jsonify({"erro": "Erro interno do servidor", "detalhe": str(e)}), 500


# ── Internals ─────────────────────────────────────────────────────────────────

def _parse_spreadsheet(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        raise ValueError("Planilha vazia")

    mapped_headers = [_map_header(h) for h in raw_headers]

    present  = {k for k in mapped_headers if k}
    required = {"numero_processo", "data_audiencia"}
    missing  = required - present
    if missing:
        raise ValueError(
            f"Coluna(s) obrigatória(s) não encontrada(s): {', '.join(sorted(missing))}. "
            "Esperado: 'Número do Processo' e 'Data da Audiência'."
        )

    records = []
    vazias  = 0
    for row in rows_iter:
        data = {k: v for k, v in zip(mapped_headers, row) if k}
        if not any(v for v in data.values() if v is not None and str(v).strip()):
            vazias += 1
            if vazias >= 3:
                break
            continue
        vazias = 0
        if not data.get("numero_processo"):
            continue

        # Normaliza data → DD/MM/YYYY
        raw_date = data.get("data_audiencia")
        if hasattr(raw_date, "strftime"):
            data["data_audiencia"] = raw_date.strftime("%d/%m/%Y")
        else:
            data["data_audiencia"] = str(raw_date or "").strip()

        # Normaliza horário → HH:MM
        raw_time = data.get("horario_audiencia")
        if hasattr(raw_time, "strftime"):
            data["horario_audiencia"] = raw_time.strftime("%H:%M")
        elif hasattr(raw_time, 'seconds'):  # timedelta
            h, m = divmod(raw_time.seconds // 60, 60)
            data["horario_audiencia"] = f"{h:02d}:{m:02d}"
        else:
            t = str(raw_time or "").strip()
            data["horario_audiencia"] = t[:5] if t else ""

        records.append(data)

    wb.close()
    if not records:
        raise ValueError("Nenhuma linha válida encontrada na planilha")
    return records


def _log(msg: str, status: str | None = None):
    _log_queue.put(json.dumps({"msg": msg, "status": status}))


def _run(rows: list[dict], report_path: Path):
    try:
        rpa_module.run_automation(rows, _log, report_path, _state)
        _state["report_ready"] = True
    except Exception:
        _log(f"Erro crítico:\n{traceback.format_exc()}", "error")
    finally:
        _state["running"] = False
        _state["done"]    = True
        _log_queue.put(json.dumps({"done": True}))


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Relatório Audiência RPA → http://localhost:5002")
    app.run(host="127.0.0.1", port=5002, debug=False, threaded=True)
